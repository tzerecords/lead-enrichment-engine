"""Script to create M3 test Excel file with 20 rows of test data."""

import sys
from pathlib import Path

# Add project root to Python path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def create_m3_test_excel():
    """Create Excel file with 20 rows of test data for M3 testing."""
    
    # Create DataFrame with 20 rows - ensure all lists have exactly 20 elements
    data = {
        'CIF': [
            'A12345678', 'B87654321', 'C11111111', 'D22222222', 'E33333333',
            'F44444444', 'G55555555', 'H66666666', 'I77777777', 'J88888888',
            'K99999999', 'L00000000', 'M11111111', 'N22222222', 'O33333333',
            'P44444444', 'Q55555555', 'R66666666', 'S77777777', 'T88888888'
        ],
        'RAZON_SOCIAL': [
            'Empresa A SL', 'Empresa B SA', 'Empresa C SLU', 'Empresa D SL', 'Empresa E SA',
            'Empresa F SL', 'Empresa G SA', 'Empresa H SL', 'Empresa I SA', 'Empresa J SL',
            'Empresa K SL', 'Empresa L SA', 'Empresa M SL', 'Empresa N SA', 'Empresa O SL',
            'Empresa P SA', 'Empresa Q SL', 'Empresa R SA', 'Empresa S SL', 'Empresa T SA'
        ],
        'TELEFONO': [
            '612345678', '12345', '987654321', '5555', '611222333',
            '922334455', '99', '633445566', '', '644556677',
            '655667788', '666778899', '677889900', '688990011', '699001122',
            '700112233', '711223344', '722334455', '733445566', '744556677'
        ],
        'EMAIL': [
            'valid@example.com', 'invalid.email', '', 'test@', 'good@test.com',
            'another@valid.com', 'bad@', '', 'ok@mail.es', 'final@company.com',
            'contact@empresa.com', 'info@test.es', 'ventas@company.com', 'invalid-email', 'sales@test.com',
            'comercial@empresa.es', 'noreply@test.com', 'support@company.es', '', 'admin@test.com'
        ],
        'CONSUMO_MWH': [
            100, 200, 50, 300, 150,
            80, 20, 250, 180, 90,
            120, 220, 60, 280, 140,
            75, 25, 260, 190, 95
        ],
        'LUZ': [
            'Sí', 'Sí', 'No', 'Sí', 'Sí',
            'No', 'No', 'Sí', 'Sí', 'No',
            'Sí', 'Sí', 'No', 'Sí', 'Sí',
            'No', 'No', 'Sí', 'Sí', 'No'
        ],
        'GAS': [
            'No', 'Sí', 'No', 'Sí', 'No',
            'Sí', 'No', 'Sí', 'No', 'Sí',
            'No', 'Sí', 'No', 'Sí', 'No',
            'Sí', 'No', 'Sí', 'No', 'Sí'
        ],
        'WEBSITE': [
            '', 'https://existing.com', '', '', 'https://another.com',
            '', '', 'https://third.com', '', '',
            'https://fourth.com', '', 'https://fifth.com', '', '',
            '', 'https://sixth.com', '', '', ''
        ],
        'CNAE': [
            '', '1234', '', '', '5678',
            '', '', '9012', '', '',
            '3456', '', '7890', '', '2345',
            '', '6789', '', '0123', ''
        ],
        'OBSERVACIONES': [
            'Nota inicial para empresa A', 'Observación importante empresa B', 
            'Comentario empresa C', 'Nota empresa D', 'Observación empresa E',
            'Comentario empresa F', 'Nota empresa G', 'Observación empresa H',
            'Comentario empresa I', 'Nota empresa J',
            'Observación empresa K', 'Comentario empresa L', 'Nota empresa M',
            'Observación empresa N', 'Comentario empresa O',
            'Nota empresa P', 'Observación empresa Q', 'Comentario empresa R',
            'Nota empresa S', 'Observación empresa T'
        ]
    }
    
    # Verify all lists have the same length
    lengths = {k: len(v) for k, v in data.items()}
    if len(set(lengths.values())) > 1:
        raise ValueError(f"Inconsistent data lengths: {lengths}")
    
    df = pd.DataFrame(data)
    
    # Create output path
    output_path = Path(__file__).parent / 'm3_test_data.xlsx'
    
    # Create Excel file with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Format empty strings as None for better Excel handling
            if value == '':
                cell.value = None
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(header))
        for row_idx in range(2, len(df) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Save file
    wb.save(output_path)
    print(f"✅ Excel file created: {output_path}")
    print(f"   Rows: {len(df)}")
    print(f"   Columns: {len(df.columns)}")
    print(f"\n📋 Columns:")
    for col in df.columns:
        non_empty = df[col].notna().sum() if df[col].dtype == 'object' else len(df)
        print(f"   - {col}: {non_empty} non-empty values")
    
    return output_path


if __name__ == "__main__":
    create_m3_test_excel()
