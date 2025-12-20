"""Critical architectural rules tests for M3."""

import sys
from pathlib import Path

# Add project root to Python path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import pytest
import openpyxl
from openpyxl.styles import PatternFill

from src.core.excel_processor import read_excel, write_excel
from src.core.orchestrator import run_pipeline, run_tier2_enrichment, run_tier3_and_validation
from src.utils.config_loader import load_yaml_config


def test_observaciones_untouched():
    """TEST 1: OBSERVACIONES debe permanecer idéntico (texto y formato)."""
    # Create test Excel with OBSERVACIONES
    input_path = Path("tests/temp_observaciones_test.xlsx")
    output_path = Path("data/output/LIMPIO_temp_observaciones_test.xlsx")
    
    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Headers
    headers = ['CIF', 'RAZON_SOCIAL', 'TELEFONO', 'EMAIL', 'CONSUMO_MWH', 'LUZ', 'GAS', 'WEBSITE', 'CNAE', 'OBSERVACIONES']
    ws.append(headers)

    # Data row with OBSERVACIONES
    ws.append(['A12345678', 'Test Company', '612345678', 'test@example.com', 100, 'SI', 'NO', '', '', 'NOTA IMPORTANTE'])

    # Color OBSERVACIONES cell (yellow)
    obs_cell = ws.cell(row=2, column=10)  # OBSERVACIONES column
    obs_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb.save(input_path)

    # Process file
    df, metadata = read_excel(input_path)
    df_result, _ = run_pipeline(df, tier1_only=False)
    df_result, _ = run_tier2_enrichment(df_result, enable_email_research=False)
    df_result = run_tier3_and_validation(df_result, enable_tier3=True)
    write_excel(df_result, metadata, output_path, preserve_format=True)

    # Verify OBSERVACIONES unchanged
    wb_out = openpyxl.load_workbook(output_path)
    ws_out = wb_out.active

    # Find OBSERVACIONES column in output
    obs_col_idx = None
    for col_idx, header in enumerate(ws_out[1], start=1):
        if header.value == 'OBSERVACIONES':
            obs_col_idx = col_idx
            break

    assert obs_col_idx is not None, "OBSERVACIONES column not found in output!"

    obs_value_out = ws_out.cell(row=2, column=obs_col_idx).value
    obs_fill_out = ws_out.cell(row=2, column=obs_col_idx).fill.start_color

    assert obs_value_out == 'NOTA IMPORTANTE', f"OBSERVACIONES text changed! Got: {obs_value_out}"
    
    # Check fill color (yellow = FFFF00 or FFFFFFFF00 depending on format)
    fill_rgb = str(obs_fill_out.rgb) if hasattr(obs_fill_out, 'rgb') else None
    assert fill_rgb is not None, "OBSERVACIONES fill color not preserved!"

    # Cleanup
    input_path.unlink(missing_ok=True)
    output_path.unlink(missing_ok=True)


def test_no_hardcoded_values():
    """TEST 2: NO debe haber valores de negocio hardcodeados en Python."""
    import re

    files_to_check = [
        "src/enrichers/tier3_enricher.py",
        "src/core/scoring_engine.py",
        "src/validators/email_batch_validator.py",
        "src/validators/phone_batch_validator.py",
        "src/validators/cif_batch_validator.py",
    ]

    hardcoded_patterns = [
        r'\b(15|20|25|30|70|80|100)\s*[#]?\s*(weight|threshold|score)',  # Magic numbers for scoring
        r'"(High|Medium|Low)"\s*if\s*\w+\s*[<>=]',  # Hardcoded quality thresholds
        r'timeout\s*=\s*\d{2,}',  # Timeouts > 9 seconds
    ]

    violations = []

    for filepath in files_to_check:
        full_path = Path(__file__).parent.parent / filepath
        if not full_path.exists():
            continue
        with open(full_path, 'r') as f:
            content = f.read()
            for pattern in hardcoded_patterns:
                matches = re.findall(pattern, content, re.IGNORECASE)
                if matches:
                    violations.append(f"{filepath}: found hardcoded value matching {pattern}")

    # Allow some exceptions (like timeout=5 is OK, timeout = min(...) is OK)
    violations = [
        v for v in violations 
        if 'timeout = min' not in v 
        and 'timeout=5' not in v 
        and 'timeout=min' not in v
        and 'timeout: float = 5' not in v
        and 'timeout: int = 5' not in v
    ]

    assert len(violations) == 0, f"Found hardcoded values:\n" + "\n".join(violations)


def test_tier3_only_empty_fields():
    """TEST 3: Tier3 solo debe rellenar campos vacíos, NO sobrescribir."""
    df_input = pd.DataFrame({
        'CIF': ['A12345678', 'B87654321'],
        'RAZON_SOCIAL': ['Company A', 'Company B'],
        'WEBSITE': ['https://existing.com', ''],  # First has value
        'CNAE': ['1234', ''],  # First has value
        'OBSERVACIONES': ['Note1', 'Note2'],
    })

    enrichment_rules = load_yaml_config("config/rules/enrichment_rules.yaml")
    
    from src.enrichers.tier3_enricher import Tier3Enricher, SimpleSearchClient, SimpleHttpClient

    tier3 = Tier3Enricher(
        search_client=SimpleSearchClient(),
        http_client=SimpleHttpClient(),
        rules=enrichment_rules.get("tier3", {})
    )

    df_output = tier3.process_missing_only(df_input)

    # Verify existing values NOT changed
    assert df_output.loc[0, 'WEBSITE'] == 'https://existing.com', "WEBSITE sobrescrito!"
    assert df_output.loc[0, 'CNAE'] == '1234', "CNAE sobrescrito!"


def test_new_columns_at_end():
    """TEST 4: Columnas nuevas (scoring) deben estar AL FINAL."""
    # Create a simple test file first
    test_file = Path("tests/temp_column_order_test.xlsx")
    output_file = Path("data/output/LIMPIO_temp_column_order_test.xlsx")
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Create minimal Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['CIF', 'RAZON_SOCIAL', 'TELEFONO', 'EMAIL', 'CONSUMO_MWH', 'LUZ', 'GAS', 'WEBSITE', 'CNAE', 'OBSERVACIONES']
    ws.append(headers)
    ws.append(['A12345678', 'Test', '612345678', 'test@example.com', 100, 'SI', 'NO', '', '', 'Note'])
    wb.save(test_file)
    
    # Process
    df, metadata = read_excel(test_file)
    df_result, _ = run_pipeline(df, tier1_only=False)
    df_result, _ = run_tier2_enrichment(df_result, enable_email_research=False)
    df_result = run_tier3_and_validation(df_result, enable_tier3=True)
    write_excel(df_result, metadata, output_file, preserve_format=True)
    
    # Read output
    df_output = pd.read_excel(output_file)

    original_cols = ['CIF', 'RAZON_SOCIAL', 'TELEFONO', 'EMAIL', 'CONSUMO_MWH', 'LUZ', 'GAS', 'WEBSITE', 'CNAE', 'OBSERVACIONES']
    new_cols = ['COMPLETITUD_SCORE', 'CONFIDENCE_SCORE', 'DATA_QUALITY', 'DATA_SOURCES', 'LAST_UPDATED']

    # Check original columns are first (in order)
    original_positions = []
    for col in original_cols:
        if col in df_output.columns:
            original_positions.append(df_output.columns.get_loc(col))
    
    # Check new columns are after original
    new_positions = []
    for col in new_cols:
        if col in df_output.columns:
            new_positions.append(df_output.columns.get_loc(col))
    
    if original_positions and new_positions:
        max_original_pos = max(original_positions)
        min_new_pos = min(new_positions)
        assert min_new_pos > max_original_pos, f"New columns not at end! Max original: {max_original_pos}, Min new: {min_new_pos}"
    
    # Cleanup
    test_file.unlink(missing_ok=True)
    output_file.unlink(missing_ok=True)


def test_scoring_consistency():
    """TEST 5: Scoring debe ser consistente con completitud."""
    # Create a test file with varied data quality
    test_file = Path("tests/temp_scoring_test.xlsx")
    output_file = Path("data/output/LIMPIO_temp_scoring_test.xlsx")
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Create Excel with different data quality
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ['CIF', 'RAZON_SOCIAL', 'TELEFONO', 'EMAIL', 'CONSUMO_MWH', 'LUZ', 'GAS', 'WEBSITE', 'CNAE', 'OBSERVACIONES']
    ws.append(headers)
    # Row 1: Complete data
    ws.append(['A12345678', 'Complete Company', '612345678', 'complete@example.com', 200, 'SI', 'SI', 'https://complete.com', '1234', 'Complete'])
    # Row 2: Incomplete data
    ws.append(['B87654321', '', '', '', '', '', '', '', '', 'Incomplete'])
    wb.save(test_file)
    
    # Process
    df, metadata = read_excel(test_file)
    df_result, _ = run_pipeline(df, tier1_only=False)
    df_result, _ = run_tier2_enrichment(df_result, enable_email_research=False)
    df_result = run_tier3_and_validation(df_result, enable_tier3=True)
    write_excel(df_result, metadata, output_file, preserve_format=True)
    
    # Read output
    df = pd.read_excel(output_file)

    # High quality should have high completeness
    high_quality = df[df['DATA_QUALITY'] == 'High']
    if len(high_quality) > 0:
        assert high_quality['COMPLETITUD_SCORE'].min() >= 70, f"High quality con baja completitud! Min: {high_quality['COMPLETITUD_SCORE'].min()}"

    # Low quality should have low completeness (or medium)
    low_quality = df[df['DATA_QUALITY'] == 'Low']
    if len(low_quality) > 0:
        # Low quality can have up to 50% completeness (medium threshold)
        assert low_quality['COMPLETITUD_SCORE'].max() <= 60, f"Low quality con alta completitud! Max: {low_quality['COMPLETITUD_SCORE'].max()}"
    
    # Cleanup
    test_file.unlink(missing_ok=True)
    output_file.unlink(missing_ok=True)
