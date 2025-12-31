"""Excel file processor with format preservation."""

import logging
from pathlib import Path
from typing import Tuple, Dict, Any, List
from copy import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import re

logger = logging.getLogger("lead_enrichment")


def _is_red_color(color_value: Any) -> bool:
    """Check if color is red (any shade).

    Args:
        color_value: Color value from openpyxl (can be RGB, hex, or theme).

    Returns:
        True if color appears to be red.
    """
    if color_value is None:
        return False

    # Handle RGB tuple
    if isinstance(color_value, tuple):
        r, g, b = color_value[:3]
        # Red if R is significantly higher than G and B
        return r > 150 and r > g + 50 and r > b + 50

    # Handle hex string
    if isinstance(color_value, str):
        # Remove # if present
        hex_color = color_value.replace("#", "").upper()
        # Handle ARGB format (8 chars) - remove alpha channel
        if len(hex_color) == 8:
            hex_color = hex_color[2:]  # Remove alpha (FF)
        if len(hex_color) == 6:
            try:
                r = int(hex_color[0:2], 16)
                g = int(hex_color[2:4], 16)
                b = int(hex_color[4:6], 16)
                return r > 150 and r > g + 50 and r > b + 50
            except ValueError:
                return False

    # Handle openpyxl color objects
    if hasattr(color_value, "rgb"):
        rgb_str = str(color_value.rgb)
        if rgb_str:
            # Format: "FFRRGGBB" (ARGB) or "RRGGBB" or "#RRGGBB"
            rgb_clean = rgb_str.replace("#", "").upper()
            # Remove alpha channel if present (first 2 chars if length is 8)
            if len(rgb_clean) == 8:
                rgb_clean = rgb_clean[2:]  # Remove alpha (FF)
            if len(rgb_clean) == 6:
                try:
                    r = int(rgb_clean[0:2], 16)
                    g = int(rgb_clean[2:4], 16)
                    b = int(rgb_clean[4:6], 16)
                    return r > 150 and r > g + 50 and r > b + 50
                except ValueError:
                    return False

    return False


def _detect_red_rows(workbook_path: Path) -> List[int]:
    """Detect rows with red background color.

    Args:
        workbook_path: Path to Excel file.

    Returns:
        List of row indices (1-based) that have red background.
    """
    logger.info(f"Detecting red rows in {workbook_path}")
    red_rows = []

    try:
        wb = load_workbook(workbook_path, data_only=False)
        ws = wb.active

        # Iterate through rows (skip header row 1)
        for row_idx in range(2, ws.max_row + 1):
            row_has_red = False

            # Check each cell in the row
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Check if cell has a fill with a pattern (not just default fill)
                if cell.fill and cell.fill.patternType and cell.fill.patternType != "none":
                    # Check fill color
                    fill_color = None
                    if hasattr(cell.fill, "start_color") and cell.fill.start_color:
                        fill_color = cell.fill.start_color
                    elif hasattr(cell.fill, "fgColor") and cell.fill.fgColor:
                        fill_color = cell.fill.fgColor

                    if fill_color:
                        # Try different ways to get color value
                        color_value = None
                        if hasattr(fill_color, "rgb"):
                            color_value = fill_color.rgb
                        elif hasattr(fill_color, "value"):
                            color_value = fill_color.value
                        elif hasattr(fill_color, "index"):
                            # Indexed color
                            pass

                        if color_value and _is_red_color(color_value):
                            row_has_red = True
                            logger.debug(f"Row {row_idx} cell {col_idx} is red: {color_value}")
                            break

            if row_has_red:
                red_rows.append(row_idx)
                logger.debug(f"Row {row_idx} detected as red")

        wb.close()
        logger.info(f"Found {len(red_rows)} red rows: {red_rows}")

    except Exception as e:
        logger.error(f"Error detecting red rows: {e}")
        raise

    return red_rows


def read_excel(filepath: Path) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """Read Excel file and detect red rows.

    Args:
        filepath: Path to input Excel file.

    Returns:
        Tuple of (DataFrame, metadata dict).
        Metadata includes: red_row_indices (1-based), original_columns, etc.
    """
    logger.info(f"Reading Excel file: {filepath}")

    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    # Detect red rows first
    red_row_indices = _detect_red_rows(filepath)

    # Read with pandas
    df = pd.read_excel(filepath, engine="openpyxl")

    # Mark red rows (add column to indicate red status)
    # Convert 1-based row indices to 0-based DataFrame indices
    # Row 2 in Excel = index 1 in DataFrame (row 1 is header)
    red_df_indices = [idx - 2 for idx in red_row_indices if idx >= 2]
    df["_IS_RED_ROW"] = False
    df.loc[red_df_indices, "_IS_RED_ROW"] = True

    logger.info(f"Read {len(df)} rows, {len(red_df_indices)} marked as red")

    # Ensure filepath is a Path object for consistency
    filepath_obj = Path(filepath) if not isinstance(filepath, Path) else filepath
    
    metadata = {
        "red_row_indices": red_row_indices,
        "red_df_indices": red_df_indices,
        "original_columns": list(df.columns),
        "filepath": str(filepath_obj),  # Store as string for JSON serialization compatibility
    }
    
    logger.debug(f"Created metadata with filepath: {metadata['filepath']}")
    logger.debug(f"Filepath exists: {filepath_obj.exists()}")
    
    # Debug logging for metadata (workbook_context equivalent)
    logger.info(f"DEBUG read_excel: returning metadata with keys: {metadata.keys() if metadata else 'None'}")
    logger.info(f"DEBUG read_excel: metadata type: {type(metadata)}")
    logger.info(f"DEBUG read_excel: metadata is None? {metadata is None}")
    if metadata is not None:
        logger.info(f"DEBUG read_excel: metadata keys: {metadata.keys() if isinstance(metadata, dict) else 'NOT A DICT'}")

    return df, metadata


def _auto_adjust_column_widths(ws) -> None:
    """Auto-adjust column widths for a worksheet.
    
    Args:
        ws: openpyxl worksheet object.
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def write_excel(
    df: pd.DataFrame,
    metadata: Dict[str, Any],
    output_path: Path,
    preserve_format: bool = True,
    force_tier2: bool = False,
) -> None:
    """Write DataFrame to Excel with 3 sheets: BBDD ORIGINAL, HIGHLIGHT, DATOS_TÉCNICOS.

    Args:
        df: DataFrame to write.
        metadata: Metadata from read_excel (includes original filepath).
        output_path: Path for output Excel file.
        preserve_format: Whether to preserve original formatting (only for HOJA 1).
    """
    logger.info(f"Writing Excel file with 3 sheets: {output_path}")

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Get red row indices from metadata
    red_df_indices = metadata.get("red_df_indices", [])
    
    # Prepare data: remove red rows for HOJA 1 and HOJA 2, keep all for HOJA 3
    if "_IS_RED_ROW" in df.columns:
        df_processed = df[df["_IS_RED_ROW"] == False].copy()
        df_processed = df_processed.drop(columns=["_IS_RED_ROW"])
        # Keep all rows (including red) for HOJA 3
        df_all = df.drop(columns=["_IS_RED_ROW"]).copy()
    else:
        df_processed = df.copy()
        df_all = df.copy()
    
    # ============================================
    # Calculate COLOR column - SIMPLE colors
    # ============================================
    def calculate_color_status(row_idx: int, row: pd.Series) -> str:
        """Calculate COLOR status for a row - simplified colors.
        
        Args:
            row_idx: DataFrame index (0-based).
            row: Series representing the row.
            
        Returns:
            Color status string: VERDE, AMARILLO, GRIS, or ROJO
        """
        # Check if row was originally red (ignored)
        if row_idx in red_df_indices:
            return "ROJO"
        
        # Get priority
        priority = row.get("PRIORITY")
        if pd.isna(priority) or priority is None:
            priority = 0
        else:
            priority = float(priority)
        
        # Check if has new data (phone from google_places/tavily, or email found)
        has_new_phone = (
            pd.notna(row.get("PHONE")) and 
            str(row.get("PHONE_SOURCE", "")).strip() in ["google_places", "tavily"]
        )
        has_new_email = (
            pd.notna(row.get("EMAIL_SPECIFIC")) and 
            str(row.get("EMAIL_SPECIFIC", "")).strip() not in ["", "NO_EMAIL_FOUND", "NOT_FOUND"]
        )
        has_new_data = has_new_phone or has_new_email
        
        # Priority >= 2: VERDE if has new data, AMARILLO if not
        if priority >= 2:
            return "VERDE" if has_new_data else "AMARILLO"
        
        # Priority < 2: GRIS
        return "GRIS"
    
    # Calculate COLOR for all rows in df_all (includes red rows)
    df_all["COLOR"] = df_all.apply(
        lambda row: calculate_color_status(row.name, row),
        axis=1
    )
    
    # Also add to df_processed for HOJA 1
    df_processed["COLOR"] = df_processed.apply(
        lambda row: calculate_color_status(row.name, row),
        axis=1
    )

    # ============================================
    # HOJA 1: "BBDD ORIGINAL" - Columnas originales + 6 nuevas
    # ============================================
    original_columns = metadata.get("original_columns", [])
    # Remove _IS_RED_ROW from original columns if present
    original_columns = [col for col in original_columns if col != "_IS_RED_ROW"]
    
    # Add 7 new columns (including COLOR)
    new_columns = ['PRIORITY', 'WEBSITE', 'CNAE', 'EMAIL_SPECIFIC', 'CONTACT_NAME', 'DATA_QUALITY', 'COLOR']
    
    # Build list of columns for HOJA 1
    hoja1_columns = []
    for col in original_columns:
        if col in df_processed.columns:
            hoja1_columns.append(col)
    for col in new_columns:
        if col in df_processed.columns:
            hoja1_columns.append(col)
    
    # Ensure COLOR is included if it exists
    if "COLOR" in df_processed.columns and "COLOR" not in hoja1_columns:
        hoja1_columns.append("COLOR")
    
    df_original = df_processed[[col for col in hoja1_columns if col in df_processed.columns]].copy()
    
    # ============================================
    # HOJA 2: "LEADS ENRIQUECIDOS" - Todos los prioritarios (PRIORITY >= 2)
    # ============================================
    # Filter: ALL leads with PRIORITY >= 2 (all prioritized leads)
    mask_highlight = pd.Series([True] * len(df_processed), index=df_processed.index)
    
    if 'PRIORITY' in df_processed.columns:
        mask_highlight = mask_highlight & (df_processed['PRIORITY'] >= 2)
    
    logger.info(f"LEADS ENRIQUECIDOS: Filtered to {mask_highlight.sum()}/{len(mask_highlight)} rows (PRIORITY>=2)")
    
    df_highlight = df_processed[mask_highlight].copy()
    
    # If HIGHLIGHT is empty, fallback to all processed rows (shouldn't happen with PRIORITY>=2)
    if len(df_highlight) == 0:
        logger.warning("LEADS ENRIQUECIDOS is empty, falling back to all processed rows")
        df_highlight = df_processed.copy()
    
    # Create helper columns for HIGHLIGHT
    # TEL_NUEVO: PHONE formateado si existe y es de google_places/tavily
    if 'PHONE' in df_highlight.columns and 'PHONE_SOURCE' in df_highlight.columns:
        df_highlight['📞 TEL_NUEVO'] = df_highlight.apply(
            lambda row: row['PHONE'] if (
                pd.notna(row.get('PHONE')) and 
                str(row.get('PHONE_SOURCE', '')).strip() in ['google_places', 'tavily']
            ) else '',
            axis=1
        )
        
        # FIX 1: Formatear TEL_NUEVO como string para evitar notación científica
        def format_phone(phone):
            if pd.isna(phone) or phone == '':
                return ''
            phone_str = str(phone).strip()
            # Remover +34 si existe
            if phone_str.startswith('+34'):
                phone_str = phone_str[3:].strip()
            elif phone_str.startswith('34'):
                phone_str = phone_str[2:].strip()
            # Asegurar que es string y formatear
            if len(phone_str) == 9:
                return f"+34 {phone_str[0:3]} {phone_str[3:6]} {phone_str[6:]}"
            return phone_str
        
        df_highlight['📞 TEL_NUEVO'] = df_highlight['📞 TEL_NUEVO'].apply(format_phone)
    else:
        df_highlight['📞 TEL_NUEVO'] = ''
    
    # EMAIL_NUEVO: EMAIL_SPECIFIC o EMAIL_FOUND si existe y no es NO_EMAIL_FOUND
    df_highlight['📧 EMAIL_NUEVO'] = ''
    if 'EMAIL_SPECIFIC' in df_highlight.columns:
        df_highlight['📧 EMAIL_NUEVO'] = df_highlight.apply(
            lambda row: row['EMAIL_SPECIFIC'] if (
                pd.notna(row.get('EMAIL_SPECIFIC')) and 
                str(row.get('EMAIL_SPECIFIC', '')).strip() not in ['', 'NO_EMAIL_FOUND', 'NOT_FOUND']
            ) else '',
            axis=1
        )
    # También usar EMAIL_FOUND si existe (de Tavily complementary search)
    if 'EMAIL_FOUND' in df_highlight.columns:
        for idx in df_highlight.index:
            email_found = df_highlight.loc[idx, 'EMAIL_FOUND']
            email_nuevo = df_highlight.loc[idx, '📧 EMAIL_NUEVO']
            # Si EMAIL_NUEVO está vacío pero EMAIL_FOUND tiene valor, usarlo
            if (pd.isna(email_nuevo) or str(email_nuevo).strip() == '') and pd.notna(email_found) and str(email_found).strip() not in ['', 'NO_EMAIL_FOUND', 'NOT_FOUND']:
                df_highlight.loc[idx, '📧 EMAIL_NUEVO'] = email_found
    
    # RAZON_SOCIAL_NUEVA: RAZON_SOCIAL si es de google_places/tavily
    if 'RAZON_SOCIAL' in df_highlight.columns and 'RAZON_SOCIAL_SOURCE' in df_highlight.columns:
        df_highlight['🏢 RAZON_SOCIAL_NUEVA'] = df_highlight.apply(
            lambda row: row['RAZON_SOCIAL'] if (
                pd.notna(row.get('RAZON_SOCIAL')) and 
                str(row.get('RAZON_SOCIAL_SOURCE', '')).strip() in ['google_places', 'tavily']
            ) else '',
            axis=1
        )
    else:
        df_highlight['🏢 RAZON_SOCIAL_NUEVA'] = ''
    
    # RESULTADO: resumen de qué se encontró (usar columnas renombradas)
    def get_resultado(row):
        found = []
        if pd.notna(row.get('📞 TEL_NUEVO')) and str(row.get('📞 TEL_NUEVO', '')).strip():
            found.append('✅ Tel nuevo')
        if pd.notna(row.get('📧 EMAIL_NUEVO')) and str(row.get('📧 EMAIL_NUEVO', '')).strip():
            found.append('✅ Email')
        # Usar la columna renombrada si existe, sino la original
        razon_col = '🏢 RAZÓN SOCIAL' if '🏢 RAZÓN SOCIAL' in row.index else '🏢 RAZON_SOCIAL_NUEVA'
        if pd.notna(row.get(razon_col)) and str(row.get(razon_col, '')).strip():
            found.append('✅ Razón social')
        return ' | '.join(found) if found else '❌ Sin datos nuevos'
    
    df_highlight['RESULTADO'] = df_highlight.apply(get_resultado, axis=1)
    
    # FIX 2: Reorganizar columnas con estructura clara
    # Renombrar PRIORITY a PRIORIDAD con valores explicados
    if 'PRIORITY' in df_highlight.columns:
        df_highlight['PRIORIDAD'] = df_highlight['PRIORITY'].apply(
            lambda x: f"Alta ({int(x)})" if pd.notna(x) and x >= 3 else (f"Media ({int(x)})" if pd.notna(x) and x >= 2 else f"Baja ({int(x)})" if pd.notna(x) else "")
        )
    
    # Renombrar RAZON_SOCIAL_NUEVA
    if '🏢 RAZON_SOCIAL_NUEVA' in df_highlight.columns:
        df_highlight['🏢 RAZÓN SOCIAL'] = df_highlight['🏢 RAZON_SOCIAL_NUEVA']
    
    # Select highlight columns - ORGANIZED with clear structure
    highlight_cols = []
    
    # Sección 1: Información básica
    for col in ["NOMBRE CLIENTE", "CONSUMO", "PRIORIDAD"]:
        if col in df_highlight.columns:
            highlight_cols.append(col)
    
    # Sección 2: Datos nuevos encontrados
    for col in ["📞 TEL_NUEVO", "📧 EMAIL_NUEVO", "🏢 RAZÓN SOCIAL", "RESULTADO"]:
        if col in df_highlight.columns:
            highlight_cols.append(col)
    
    # Sección 3: Datos originales
    for col in ["TELEFONO 1", "MAIL ", "CIF/NIF", "CIF_NIF", "CIF", "DIRECCIÓN CLIENTE", "POBLACIÓN CLIENTE"]:
        if col in df_highlight.columns and col not in highlight_cols:
            highlight_cols.append(col)
    
    # Use available columns
    df_highlight = df_highlight[highlight_cols].copy()
    
    # FIX: Rename columns with descriptive headers explaining each output
    column_descriptions = {
        "NOMBRE CLIENTE": "NOMBRE CLIENTE",
        "CONSUMO": "CONSUMO (MWh)",
        "PRIORIDAD": "PRIORIDAD (Alta=3, Media=2, Baja=1)",
        "📞 TEL_NUEVO": "📞 TELÉFONO NUEVO (encontrado con Google/Tavily)",
        "📧 EMAIL_NUEVO": "📧 EMAIL NUEVO (encontrado con Tavily)",
        "🏢 RAZÓN SOCIAL": "🏢 RAZÓN SOCIAL NUEVA (encontrada con Google/Tavily)",
        "RESULTADO": "RESULTADO (qué datos nuevos se encontraron)",
        "TELEFONO 1": "TELEFONO 1 (original del archivo)",
        "MAIL ": "MAIL (original del archivo)",
        "CIF/NIF": "CIF/NIF (original del archivo)",
        "CIF_NIF": "CIF/NIF (original del archivo)",
        "CIF": "CIF/NIF (original del archivo)",
        "DIRECCIÓN CLIENTE": "DIRECCIÓN CLIENTE (original)",
        "POBLACIÓN CLIENTE": "POBLACIÓN CLIENTE (original)"
    }
    
    # Rename columns with descriptions
    rename_dict = {col: column_descriptions.get(col, col) for col in highlight_cols if col in column_descriptions}
    if rename_dict:
        df_highlight = df_highlight.rename(columns=rename_dict)
    
    # OBSERVACIONES: NEVER modify - must remain exactly as input
    # (removed truncation to preserve original)
    
    # Sort: PRIORITY DESC, DATA_QUALITY DESC, CONSUMO DESC
    sort_cols = []
    if 'PRIORITY' in df_highlight.columns:
        sort_cols.append('PRIORITY')
    elif 'PRIORIDAD' in df_highlight.columns:
        # Use PRIORITY for sorting if available, otherwise skip
        pass
    if 'DATA_QUALITY' in df_highlight.columns:
        sort_cols.append('DATA_QUALITY')
    if 'CONSUMO' in df_highlight.columns:
        sort_cols.append('CONSUMO')
    
    if sort_cols:
        df_highlight = df_highlight.sort_values(by=sort_cols, ascending=[False] * len(sort_cols))
    
    # ============================================
    # FIX 3: Añadir fila de resumen al inicio de LEADS ENRIQUECIDOS
    # ============================================
    total_received = len(df_all)
    red_count = len(red_df_indices)
    analyzed_count = len(df_highlight)
    # Contar leads con datos nuevos (RESULTADO no es "❌ Sin datos nuevos")
    if 'RESULTADO' in df_highlight.columns:
        with_new_data = (df_highlight['RESULTADO'] != '❌ Sin datos nuevos').sum()
    else:
        with_new_data = 0
    
    # Calcular descartadas por baja prioridad (total - red - analizadas)
    low_priority_count = total_received - red_count - analyzed_count
    
    summary_text = f"RESUMEN: {total_received} empresas recibidas | {red_count} descartadas (fila roja) | {low_priority_count} descartadas (baja prioridad) | {analyzed_count} analizadas | {with_new_data} con datos nuevos"
    
    # Crear DataFrame con resumen como primera fila (merge todas las columnas en la primera)
    # Use current column names (after renaming)
    current_cols = list(df_highlight.columns)
    summary_row = {col: summary_text if col == current_cols[0] else '' for col in current_cols}
    summary_df = pd.DataFrame([summary_row])
    df_highlight_with_summary = pd.concat([summary_df, df_highlight], ignore_index=True)
    
    # ============================================
    # Write Excel with 2 sheets (LEADS ENRIQUECIDOS first, then BBDD ORIGINAL)
    # ============================================
    # First, write all sheets using pandas - ORDER: LEADS ENRIQUECIDOS first
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_highlight_with_summary.to_excel(writer, sheet_name='LEADS ENRIQUECIDOS', index=False)
        df_original.to_excel(writer, sheet_name='BBDD ORIGINAL', index=False)
    
    # Now apply formatting
    wb = load_workbook(output_path)
    
    # ============================================
    # Format HOJA 1: Preserve original format (then apply COLOR fills)
    # ============================================
    ws_original = wb['BBDD ORIGINAL']
    
    if preserve_format and "filepath" in metadata:
        original_path = metadata["filepath"]
        original_path_obj = Path(original_path) if isinstance(original_path, str) else original_path
        
        try:
            if original_path_obj.exists():
                wb_source = load_workbook(original_path_obj)
                ws_source = wb_source.active
                
                # Copy header formatting
                for col_idx in range(1, min(len(hoja1_columns) + 1, ws_source.max_column + 1)):
                    cell = ws_original.cell(row=1, column=col_idx)
                    if col_idx <= ws_source.max_column:
                        orig_cell = ws_source.cell(row=1, column=col_idx)
                        try:
                            if orig_cell.fill and orig_cell.fill.patternType:
                                cell.fill = PatternFill(
                                    start_color=orig_cell.fill.start_color,
                                    end_color=orig_cell.fill.end_color,
                                    fill_type=orig_cell.fill.fill_type
                                )
                        except Exception:
                            pass
                        try:
                            if orig_cell.font:
                                cell.font = orig_cell.font.copy()
                        except Exception:
                            pass
                
                # Copy data row formatting (use row 2 as template) - BUT skip fill (we'll apply COLOR-based fills)
                original_columns_count = len(original_columns)
                for row_idx in range(2, ws_original.max_row + 1):
                    for col_idx in range(1, len(hoja1_columns) + 1):
                        cell = ws_original.cell(row=row_idx, column=col_idx)
                        if col_idx <= original_columns_count and ws_source.max_row >= 2:
                            try:
                                orig_cell = ws_source.cell(row=2, column=col_idx)
                                if orig_cell.has_style:
                                    cell.font = copy(orig_cell.font)
                                    # Skip fill - we'll apply COLOR-based fills below
                                    cell.border = copy(orig_cell.border)
                                    cell.alignment = copy(orig_cell.alignment)
                                    cell.number_format = orig_cell.number_format
                            except Exception:
                                pass
                
                wb_source.close()
        except Exception as e:
            logger.warning(f"Could not preserve format for HOJA 1: {e}")
    
    # Apply COLOR-based background colors AFTER preserving other formatting
    # Find COLOR column index
    color_col_idx = None
    for idx, col_name in enumerate(hoja1_columns, start=1):
        if col_name == "COLOR":
            color_col_idx = idx
            break
    
    # Define color mappings - SIMPLE colors (ARGB format for openpyxl)
    color_fills = {
        "VERDE": PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid"),  # Light green - Datos nuevos
        "AMARILLO": PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),  # Light yellow - Prioritario sin datos
        "GRIS": PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"),  # Light gray - No prioritario
        "ROJO": PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),  # Light red - Ignorado
    }
    
    # Apply background colors to entire rows based on COLOR column
    if color_col_idx:
        logger.info(f"Applying COLOR-based background fills to HOJA 1 (COLOR column at index {color_col_idx})")
        for row_idx in range(2, ws_original.max_row + 1):  # Start from row 2 (skip header)
            color_cell = ws_original.cell(row=row_idx, column=color_col_idx)
            color_value = str(color_cell.value or "").strip()
            
            if color_value in color_fills:
                fill = color_fills[color_value]
                # Apply fill to entire row
                for col_idx in range(1, ws_original.max_column + 1):
                    cell = ws_original.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
    
    # Auto-adjust widths for HOJA 1
    _auto_adjust_column_widths(ws_original)
    
    # ============================================
    # Format HOJA 2: Headers bold, clean white background, summary row
    # ============================================
    ws_highlight = wb['LEADS ENRIQUECIDOS']
    
    # FIX 3: Format summary row (row 1) - bold, background color, merged cells
    from openpyxl.styles import Alignment
    summary_fill = PatternFill(start_color="FFE6E6FA", end_color="FFE6E6FA", fill_type="solid")  # Lavender
    
    # Merge cells in row 1 for summary (span all columns)
    num_cols = len(list(df_highlight.columns))
    ws_highlight.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    summary_cell = ws_highlight.cell(row=1, column=1)
    summary_cell.fill = summary_fill
    summary_cell.font = Font(bold=True, size=11)
    summary_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Make headers bold (row 2, since row 1 is summary) and add background color
    header_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")  # Light gray
    for cell in ws_highlight[2]:
        cell.font = Font(bold=True, size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-adjust widths
    _auto_adjust_column_widths(ws_highlight)
    
    # Save workbook (only 2 sheets now: LEADS ENRIQUECIDOS and BBDD ORIGINAL)
    wb.save(output_path)
    wb.close()
    
    logger.info(f"Excel file written with 2 sheets (LEADS ENRIQUECIDOS, BBDD ORIGINAL): {output_path}")

